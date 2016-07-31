# -*- coding: utf-8 -*-
from openpyxl import load_workbook

def read_data(filename,n):
    inwb = load_workbook(filename=filename)
    sheetName = inwb.get_sheet_names()[n]
    sheet = inwb[sheetName]
    rows = len(sheet.rows)
    columns = len(sheet.columns)
    data = []
    for i in range(2, rows + 1, 1):
        data2 = []
        for j in range(1, columns + 1, 1):
            data2.append(sheet.cell(row=i, column=j).value)
        data.append(data2)
    return rows, columns, data
